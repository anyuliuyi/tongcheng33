import os

from openpyxl import Workbook,load_workbook
import threading
from framework.httprequests import httprequest
from framework.framework import gl
from openpyxl.styles import Font, colors, Alignment
import time
from decimal import Decimal


excelpath=os.getcwd()+'/test-data/intent_online.xlsx'
testresult=os.getcwd()+'/test-results/stdfaq-results.xlsx'
url='http://poc1.emotibot.com:1680/v1/openapi'      #设置接口url

passed = 0
failed = 0

if os.path.exists(testresult):
    os.remove(testresult)

wb = load_workbook(excelpath)
ws = wb['Sheet1']
excelSize = ws.max_row    #获取列数
print(excelSize)

print('\n\n********************************* 测试开始 *********************************')
# 判断单元格不为空，则 打印出 单元格中的值
class IntenrtQbean:
    question = ''
    answer = ''
    testresult = ''
    remarks = ''
# 测试问题对象list
questionsCollection=[]
# 测试结果问题list
questionsCollectionExcel=[]

for i in range(2, excelSize + 1):
    testsetTemp = IntenrtQbean();
    IntenrtQbean.question = ws.cell(i, 1).value
    IntenrtQbean.answer = ws.cell(i, 2).value
    questionsCollection.append(IntenrtQbean)
    print('工作表：'+str(i-1)," 测试问题>>>>>>      ",IntenrtQbean.question," | ","测试意图>>>>>>      ",IntenrtQbean.answer)
print('>>>>>> 读取excel测试数据结束 >>>>>>')


rlock = threading.RLock()

class cxThreadWorker(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
    def run(self):
        global questionsCollection
        global rlock
        global passed,failed

        while len(questionsCollection) > 0:
            # 得到 第一个对象，下标为 0
            IntenrtQbean =questionsCollection.__getitem__(0)
            intentQ = IntenrtQbean.question
            intentA = IntenrtQbean.answer

            # 删除掉 下标为 0 第一个对象，须要 加锁
            rlock.acquire()
            try:
                questionsCollection.pop(0)
                print('共：' + str(excelSize - 1) + ' 个, 当前：' + str(excelSize - 1 - len(questionsCollection)))
            except Exception:
                break
            finally:
                rlock.release()

            # 头信息
            header = {'Content-Type': 'application/json', 'sessionId': str(gl.getTimestamp()),
                          'userId': str(gl.getTimestamp()), 'appId': 'b6ad1f823d8d4ffea36568330239f237'}

            # body（JSON） 信息
            bodystringStdQ = '{ "text": "' + intentQ + '"}'


            resposeJsonStdQ = httprequest.sendPostwithHeaders(url, header, bodystringStdQ)
            flag = False
            try:
                actualAnswer =resposeJsonStdQ['info']['intent']
            except BaseException:
                flag = True

            if flag==False:
                if gl.repaceSpecialCharactersinString(actualAnswer) == gl.repaceSpecialCharactersinString(intentA):
                    IntenrtQbean.testresult == 'pass'
                else:
                    IntenrtQbean.testresult == 'fail'
                    remarks = '>>>期望意图：\n' + intentA + '\n>>>实际返回答案：\n' + actualAnswer
                    IntenrtQbean.remarks = remarks
            else:
                IntenrtQbean.testresult == 'fail'
                remarks = '>>>接口返回结果错误，见：\n' + str(resposeJsonStdQ)
                IntenrtQbean.remarks = remarks

            questionsCollectionExcel.append(IntenrtQbean)  #将测试结果写入测试集


now=lambda :time.time()
start=now()

subThreads=[]

multiThreadCount=10 # set threads to ran the test
for r in range(multiThreadCount):
    thread = cxThreadWorker()
    thread.start()
    subThreads.append(thread)

for tt in subThreads:
    tt.join()

# ******************************************** 写入结果到excel ********************************************
wb=Workbook()
ws=wb.active
# 表头
ws.column_dimensions['A'].width = 50    #测试问题
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 120
ws.append(['标准问题','测试结果','备注'])
for r in range(len(questionsCollectionExcel)):
    stdQuestion=questionsCollectionExcel.__getitem__(r).question
    result=questionsCollectionExcel.__getitem__(r).testresult
    remarks=questionsCollectionExcel.__getitem__(r).remarks
    bold_itatic_24_font_fail = Font(name='Calibri', size=11, italic=False, color=colors.RED, bold=True)
    bold_itatic_24_font_pass = Font(name='Calibri', size=11, italic=False, color=colors.GREEN, bold=True)

    alignment = Alignment(wrap_text=True)   # 设置 备注单元格为自动换行

    ws.append([stdQuestion, result, remarks])

    if result == 'fail':
        ws['B' + str(r + 2)].font = bold_itatic_24_font_fail
        ws['C' + str(r + 2)].alignment = alignment
        failed+=1
    else:
        ws['B' + str(r + 2)].font = bold_itatic_24_font_pass
        ws['C' + str(r + 2)].alignment = alignment
        passed+=1


wb.save(testresult)

print('\n\n>>>>>> FAQ全匹配 测试结束！ 用时：',str(int(now()-start))+' 秒')
print('>>>>>> 共：',str(passed+failed)+' 个',' 通过率:',str(Decimal((passed/(passed+failed))*100).quantize(Decimal('0.00')))+'%',' 通过：'+str(passed)+' 个，', '失败：'+str(failed)+' 个')
print('\n>>>>>> 测试结果路径：'+testresult)
print('\n\n********************************* 测试结束 *********************************')




