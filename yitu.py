#coding=utf-8

import os
from congfig import congfig
from openpyxl import Workbook,load_workbook
import threading
from framework.httprequests import httprequest
from framework.framework import gl
from openpyxl.styles import Font, colors, Alignment
import time
from decimal import Decimal

# 测试文件路径
excelpath = congfig.intent_file
# 测试结果文件路径
testresult = congfig.intent_result

# 测试接口url
url = congfig.openapi_url


# 统计测试结果参数
passed=0
failed=0


# 判断测试结果是否存在，存在就移除
if os.path.exists(testresult):
    os.remove(testresult)

# 打开测试文件，读取测试文件最大行值
wb = load_workbook(excelpath)
sheet1 = wb['Sheet1']
excelSize = sheet1.max_row
# excelSize=10

# 将每一行当成一个对象，每列就是对象的属性
class StandardQBean:
    question=''
    answer=''
    testresult=''
    remarks=''

# 测试集集合
questionsCollection = []

# 测试结果集合
questionsCollectionExcel = []

print('\n\n********************************* 测试开始 *********************************')

# 判断单元格不为空，则 打印出 单元格中的值，从第二行读取
for i in range(2,excelSize+1):
    if sheet1.cell(i, 2).value != None:
        # print(sheet1.cell(i, 5).value)
        stdQTemp = StandardQBean();
        stdQTemp.question = sheet1.cell(i, 1).value
        stdQTemp.answer = sheet1.cell(i, 2).value
        questionsCollection.append(stdQTemp)
        # print('>>>>>> ',str(i-1),'. ',stdQTemp.question)

print('>>>>>> 读取excel测试数据结束 >>>>>>')


rlock=threading.RLock()

class cxThreadWorker(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
    def run(self):
        global questionsCollection
        global rlock
        global passed,failed

        while len(questionsCollection) > 0:
            # print(threading.current_thread().getName())
            # 得到 第一个对象，下标为 0
            standardQBeanTemp = questionsCollection.__getitem__(0)
            stdQ = standardQBeanTemp.question
            stdA = standardQBeanTemp.answer

            # 删除掉 下标为 0 第一个对象，须要 加锁
            rlock.acquire()
            try:
                questionsCollection.pop(0)
                print('共：' + str(excelSize - 1) + ' 个, 当前：' + str(excelSize - 1 - len(questionsCollection)))
            except Exception:
                break
            finally:
                rlock.release()

            # 头信息
            header = {'Content-Type': 'application/json', 'sessionId': str(gl.getTimestamp()),
                          'userId': str(gl.getTimestamp()), 'appId':congfig.online_appId}
            # body（JSON） 信息
            bodystringStdQ = '{ "text": "' + stdQ + '"}'

            resposeJsonStdQ = httprequest.sendPostwithHeaders(url, header, bodystringStdQ)
            flag=False
            try:
                actualAnswer = resposeJsonStdQ['info']['intent']
                module = resposeJsonStdQ['info']['module']
            except BaseException:
                flag=True

            if flag==False:
                if gl.repaceSpecialCharactersinString(actualAnswer) == gl.repaceSpecialCharactersinString(
                        stdA) and module == "task_engine":
                    # print('pass')
                    standardQBeanTemp.testresult = 'pass'
                else:
                    standardQBeanTemp.testresult = 'fail'
                    remarks = '>>>期望标准答案：\n' + stdA + '\n>>>实际返回答案：\n' + actualAnswer
                    standardQBeanTemp.remarks = remarks
                    # print(remarks)
            else:
                standardQBeanTemp.testresult = 'fail'
                remarks = '>>>接口返回结果错误，见：\n' + str(resposeJsonStdQ)
                standardQBeanTemp.remarks = remarks

            questionsCollectionExcel.append(standardQBeanTemp)


now = lambda :time.time()
start = now()

subThreads=[]

multiThreadCount=10 # set threads to ran the test
for r in range(multiThreadCount):
    thread = cxThreadWorker()
    thread.start()
    subThreads.append(thread)

for tt in subThreads:
    tt.join()

# ******************************************** 写入结果到excel ********************************************
wb = Workbook()
ws = wb.active
# 表头
ws.column_dimensions['A'].width = 50    #测试问题
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 120
ws.append(['标准问题','测试结果','备注'])
for r in range(len(questionsCollectionExcel)):
    stdQuestion = questionsCollectionExcel.__getitem__(r).question
    result = questionsCollectionExcel.__getitem__(r).testresult
    remarks = questionsCollectionExcel.__getitem__(r).remarks
    bold_itatic_24_font_fail = Font(name='Calibri', size=11, italic=False, color=colors.RED, bold=True)
    bold_itatic_24_font_pass = Font(name='Calibri', size=11, italic=False, color=colors.GREEN, bold=True)

    alignment = Alignment(wrap_text=True)   # 设置 备注单元格为自动换行

    ws.append([stdQuestion, result, remarks])

    if result == 'fail':
        ws['B' + str(r + 2)].font = bold_itatic_24_font_fail
        ws['C' + str(r + 2)].alignment = alignment
        failed += 1
    else:
        ws['B' + str(r + 2)].font = bold_itatic_24_font_pass
        ws['C' + str(r + 2)].alignment = alignment
        passed += 1

# 保存
wb.save(testresult)

print('\n\n>>>>>> FAQ意图 测试结束！ 用时：',str(int(now()-start))+' 秒')
print('>>>>>> 共：',str(passed+failed)+' 个',' 通过率:',str(Decimal((passed/(passed+failed))*100).quantize(Decimal('0.00'))) +'%',' 通过：'+str(passed)+' 个，', '失败：'+str(failed)+' 个')
print('\n>>>>>> 测试结果路径：'+testresult)
print('\n\n********************************* 测试结束 *********************************')
