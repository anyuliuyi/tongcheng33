

import os

from openpyxl import Workbook,load_workbook
import threading
from framework.httprequests import httprequest
from framework.framework import gl
from openpyxl.styles import Font, colors, Alignment
import time
from decimal import Decimal

# 测试数据excel路径
excelpath=os.getcwd()+'/test-data/newtest1.8.xlsx'  # taibaotestcasereduce.xlsx    faqdata.xlsx
# 测试结果路径
testresult=os.getcwd()+'/test-results/accuracyrate-results.xlsx'
url='http://poc1.emotibot.com:1680/v1/openapi'

passed=0
failed=0


class TestsetBean:
    testmodule=''
    testquesiton=''
    stdquestion=''
    answer=''
    testresult=''
    remarks=''


if os.path.exists(testresult):
    os.remove(testresult)
# 测试问题对象list
questionsCollection=[]
# 测试结果问题list
questionsCollectionExcel=[]

wb=load_workbook(excelpath)
ws=wb['Sheet1']
excelSize=ws.max_row
for i in range(2, excelSize + 1):
    testsetTemp = TestsetBean();
    testsetTemp.testquestion = ws.cell(i, 1).value
    testsetTemp.stdquestion = ws.cell(i, 2).value
    questionsCollection.append(testsetTemp)
    # print('工作表：'+str(i-1)," >>>>>>      ",testsetTemp.testquestion," | ",testsetTemp.stdquestion)

# 所有测试问的个数
questioncount=len(questionsCollection)

print('\n\n********************************* 测试开始 *********************************')

print('>>>>>> 读取excel测试数据结束 >>>>>>')

rlock=threading.RLock()

class cxThreadWorker(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
    def run(self):
        global questionsCollection
        global rlock
        global passed,failed


        while len(questionsCollection) > 0:
            # print(threading.current_thread().getName())
            # 得到 第一个对象，下标为 0
            standardQBeanTemp = questionsCollection.__getitem__(0)
            StdQ = standardQBeanTemp.stdquestion
            TestQ = standardQBeanTemp.testquestion

            # 删除掉 下标为 0 第一个对象，须要 加锁
            rlock.acquire()
            try:
                questionsCollection.pop(0)
                print('共：' + str(questioncount) + ' 个, 当前：' + str(questioncount - len(questionsCollection)))
            except Exception:
                break
            finally:
                rlock.release()
            # 头信息
            header = {'Content-Type': 'application/json','sessionId': str(gl.getTimestamp()),'userId': str(gl.getTimestamp()),'appId': 'b6ad1f823d8d4ffea36568330239f237'}


            # body（JSON） 信息
            bodystringTestQ = '{ "text": "' + str(TestQ) + '"}'
            bodystringStdQ = '{ "text": "' + str(StdQ) + '"}'

            resposeJsonTestQ = httprequest.sendPostwithHeaders(url, header, bodystringTestQ)
            resposeJsonStdQ = httprequest.sendPostwithHeaders(url, header, bodystringStdQ)

            flag=False
            try:


                AnswerStdQ = resposeJsonStdQ['data'][0]['value']


            except Exception as e:
                flag=True

            if flag == False:
                if gl.repaceSpecialCharactersinString(AnswerStdQ) == gl.repaceSpecialCharactersinString(
                        AnswerStdQ):
                    standardQBeanTemp.testresult = 'pass'
                    # print('---> Pass')

                else:

                    standardQBeanTemp.testresult = 'fail'
                    if len(resposeJsonTestQ['data']) > 1:
                        remarks = '>>>标准问返回：\n' + AnswerStdQ + '\n>>>测试问返回：\n' + AnswerStdQ \
                        + '\n>>>测试相关问返回：\n' + str(resposeJsonTestQ['data'])
                        standardQBeanTemp.remarks = remarks
                    else:
                        remarks = '>>>标准问返回：\n' + AnswerStdQ + '\n>>>测试问返回：\n' + AnswerStdQ \
                        + '\n>>>测试相关问返回：\n' + str(resposeJsonTestQ['data'])
                        standardQBeanTemp.remarks = remarks

            else:
                standardQBeanTemp.testresult = 'fail'
                remarks = '>>>标准问接口返回：\n' + str(resposeJsonStdQ) + '\n>>>测试问接口返回：\n' \
                          + str(resposeJsonTestQ)

                standardQBeanTemp.remarks = remarks

            questionsCollectionExcel.append(standardQBeanTemp)


now=lambda :time.time()
start=now()

subThreads=[]

multiThreadCount=20 # set threads count for running the test
for r in range(multiThreadCount):
    thread = cxThreadWorker()
    thread.start()
    subThreads.append(thread)

for tt in subThreads:
    tt.join()


# ******************************************** 写入结果到excel ********************************************
wb=Workbook()
ws=wb.active
# 表头
ws.column_dimensions['A'].width = 50    # 测试问题
ws.column_dimensions['B'].width = 50    # 标准问题
ws.column_dimensions['C'].width = 8     # 测试结果
ws.column_dimensions['D'].width = 120   # 备注

ws.append(['测试问题','标准问题','测试结果','备注'])
for r in range(len(questionsCollectionExcel)):
    testQuestion=questionsCollectionExcel.__getitem__(r).testquestion
    stdQuestion = questionsCollectionExcel.__getitem__(r).stdquestion
    result=questionsCollectionExcel.__getitem__(r).testresult
    remarks=questionsCollectionExcel.__getitem__(r).remarks
    bold_itatic_24_font_fail = Font(name='Calibri', size=11, italic=False, color=colors.RED, bold=True)
    bold_itatic_24_font_pass = Font(name='Calibri', size=11, italic=False, color=colors.GREEN, bold=True)

    ws.append([testQuestion,stdQuestion, result, remarks])

    alignment = Alignment(wrap_text=True)   # 设置 备注单元格为自动换行

    if result == 'fail':
        ws['C' + str(r + 2)].font = bold_itatic_24_font_fail
        ws['D' + str(r + 2)].alignment = alignment
        failed += 1
    else:
        ws['C' + str(r + 2)].font = bold_itatic_24_font_pass
        ws['D' + str(r + 2)].alignment = alignment
        passed += 1

wb.save(testresult)


print('\n\n>>>>>> 测试结束！ 用时：',str(int(now()-start))+' 秒')
print('>>>>>> 共：',str(passed+failed)+' 个',' 通过率:',str(Decimal((passed/(passed+failed))*100).quantize(Decimal('0.00')))+'%',' 通过：'+str(passed)+' 个，', '失败：'+str(failed)+' 个')
print('\n>>>>>> 测试结果路径：'+testresult)
print('\n\n********************************* 测试结束 *********************************')

